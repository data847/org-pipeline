"""Render a ProfileResult as a row appended to a copy of the intake workbook."""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook

from .models import COLUMNS, ProfileResult


def append_row(
    result: ProfileResult,
    *,
    template: str | Path,
    out: str | Path,
    sheet: str = "Sheet1",
) -> Path:
    """Copy ``template`` to ``out`` (once) and append the result as a new data row.

    If ``out`` already exists it is appended to, so repeated runs accumulate rows.
    """
    template = Path(template)
    out = Path(out)
    if not out.exists():
        shutil.copy(template, out)

    wb = load_workbook(out)
    ws = wb[sheet]

    target_row = _first_empty_data_row(ws)
    row_values = result.as_row()
    for offset, value in enumerate(row_values):
        ws.cell(row=target_row, column=offset + 1, value=_excel_safe(value))

    wb.save(out)
    return out


def count_data_rows(path: str | Path, sheet: str = "Sheet1") -> int:
    """Number of populated data rows (excludes the header) in an output workbook."""
    path = Path(path)
    if not path.exists():
        return 0
    ws = load_workbook(path, read_only=True)[sheet]
    width = len(COLUMNS)
    rows = 0
    for row in ws.iter_rows(min_row=2, max_col=width, values_only=True):
        if any(v not in (None, "") for v in row):
            rows += 1
    return rows


def _first_empty_data_row(ws) -> int:
    """First row (after the header) whose mapped columns are all empty."""
    width = len(COLUMNS)
    row = 2
    while True:
        if all(ws.cell(row=row, column=c + 1).value in (None, "") for c in range(width)):
            return row
        row += 1


def _excel_safe(value):
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return value
